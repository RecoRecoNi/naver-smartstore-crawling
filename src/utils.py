import time
import os
from openpyxl import Workbook
from typing import Optional, Union, Dict, List

from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager


def get_driver(user_agent : str) -> webdriver:
    options = ChromeOptions()
    options.add_argument('user-agent=' + user_agent)
    options.add_argument('lang=ko_KR')
    options.add_argument('headless')        # chrome 안띄움
    options.add_argument("disable-gpu")
    options.add_argument("--no-sandbox")
    # chrome 드라이버 자동 설치
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
    return driver

def get_file_name() -> str:
    time_str = time.strftime("%Y%m%d-%H%M%S")
    return time_str + '.csv'

class OpenPyXL:
    @staticmethod
    def save_file(save_data : List[List[Dict[str, Union[str, int]]]] ) -> None:
        # 크롤링 결과
        results = save_data

        wb = Workbook()
        ws = wb.active
        ws.append(["상품명", "구매자 이름", "구매자 평점", "리뷰 내용", "맛 만족도", "grind", "weight"])

        row = 2

        for result in results:
            try:
                ws[f"A{row}"] = result["coffee"]
            except:
                pass
            try:
                ws[f"B{row}"] = result["user"]
            except:
                pass
            try:
                ws[f"C{row}"] = result["star_rating"]
            except:
                pass
            try:
                ws[f"D{row}"] = result["text"]
            except:
                pass
            try:
                ws[f"E{row}"] = result["satisfaction"]
            except:
                pass
            try:
                ws[f"F{row}"] = result["grind"]
            except:
                pass
            try:
                ws[f"G{row}"] = result["weight"]
            except:
                pass

            row += 1

        file_name = get_file_name()

        wb.save(os.path.join('./results', file_name))
        wb.close()

        print("Done!")
    
